"""
Vistas para generación de reportes financieros en PDF y Excel.
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.models import FinanceTransaction, FinanceCategory, EggProduction, FeedConsumption
from core.decorators import finance_write_required


@login_required
def financial_summary_view(request):
    """Vista de resumen financiero mensual con opciones de exportación."""
    today = timezone.now().date()
    
    # Obtener años disponibles en la base de datos
    years_with_data = FinanceTransaction.objects.filter(
        is_active=True
    ).dates('transaction_date', 'year')
    
    # Extraer solo los años
    available_years = sorted(
        list(set([date.year for date in years_with_data] + [today.year])),
        reverse=True
    )
    
    # Si no hay años, usar solo el actual
    if not available_years:
        available_years = [today.year]
    
    # Obtener mes y año de los parámetros o usar el actual
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    
    # Calcular rango de fechas del mes
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    # Obtener transacciones del mes
    transactions = FinanceTransaction.objects.filter(
        transaction_date__gte=start_date,
        transaction_date__lte=end_date,
        is_active=True
    ).select_related('category').order_by('transaction_date')
    
    # Calcular totales
    income_total = sum(
        t.amount_clp for t in transactions if t.category.type == 'income'
    )
    expense_total = sum(
        t.amount_clp for t in transactions if t.category.type == 'expense'
    )
    net_profit = income_total - expense_total
    
    # Calcular por categoría
    income_by_category = {}
    expense_by_category = {}
    
    for transaction in transactions:
        cat_name = transaction.category.category_name
        amount = transaction.amount_clp
        
        if transaction.category.type == 'income':
            income_by_category[cat_name] = income_by_category.get(cat_name, 0) + amount
        else:
            expense_by_category[cat_name] = expense_by_category.get(cat_name, 0) + amount
    
    # Datos de producción del mes
    production_total = EggProduction.objects.filter(
        production_date__gte=start_date,
        production_date__lte=end_date,
        is_active=True
    ).aggregate(total=Sum('quantity'))['total'] or 0
    
    # Consumo de alimento del mes
    feed_total = FeedConsumption.objects.filter(
        consumption_date__gte=start_date,
        consumption_date__lte=end_date,
        is_active=True
    ).aggregate(total=Sum('total_consumed_kg'))['total'] or 0
    
    context = {
        'month': month,
        'year': year,
        'available_years': available_years,
        'start_date': start_date,
        'end_date': end_date,
        'transactions': transactions,
        'income_total': income_total,
        'expense_total': expense_total,
        'net_profit': net_profit,
        'income_by_category': income_by_category,
        'expense_by_category': expense_by_category,
        'production_total': production_total,
        'feed_total': feed_total,
        'transaction_count': transactions.count(),
    }
    
    return render(request, 'reports/financial_summary.html', context)


@login_required
@finance_write_required
def export_financial_pdf(request):
    """Exportar resumen financiero a PDF."""
    today = timezone.now().date()
    
    # Obtener mes y año
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    
    # Calcular rango de fechas
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    # Obtener datos
    transactions = FinanceTransaction.objects.filter(
        transaction_date__gte=start_date,
        transaction_date__lte=end_date,
        is_active=True
    ).select_related('category').order_by('transaction_date')
    
    income_total = sum(t.amount_clp for t in transactions if t.category.type == 'income')
    expense_total = sum(t.amount_clp for t in transactions if t.category.type == 'expense')
    net_profit = income_total - expense_total
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="resumen_financiero_{year}_{month:02d}.pdf"'
    
    # Crear documento PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Título
    month_names = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    title = Paragraph(f"<b>Resumen Financiero</b><br/>{month_names[month]} {year}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información general
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.grey
    )
    
    info = Paragraph(f"Avícola Eugenio | Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}", subtitle_style)
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Resumen de totales
    summary_data = [
        ['RESUMEN FINANCIERO', 'MONTO'],
        ['Total Ingresos', f'${income_total:,.0f}'],
        ['Total Egresos', f'${expense_total:,.0f}'],
        ['Utilidad Neta', f'${net_profit:,.0f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    
    # Color especial para utilidad neta
    if net_profit >= 0:
        summary_table.setStyle(TableStyle([
            ('TEXTCOLOR', (1, 3), (1, 3), colors.green),
            ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ]))
    else:
        summary_table.setStyle(TableStyle([
            ('TEXTCOLOR', (1, 3), (1, 3), colors.red),
            ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Detalle de transacciones
    if transactions.exists():
        section_title = Paragraph("<b>Detalle de Transacciones</b>", styles['Heading2'])
        elements.append(section_title)
        elements.append(Spacer(1, 0.15*inch))
        
        # Tabla de transacciones
        trans_data = [['Fecha', 'Categoría', 'Tipo', 'Monto']]
        
        for t in transactions:
            trans_data.append([
                t.transaction_date.strftime('%d/%m/%Y'),
                t.category.category_name[:25],
                'Ingreso' if t.category.type == 'income' else 'Egreso',
                f'${t.amount_clp:,.0f}',
            ])
        
        trans_table = Table(trans_data, colWidths=[1.2*inch, 2.5*inch, 1.3*inch, 1.5*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(trans_table)
    
    # Generar PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


@login_required
@finance_write_required
def export_financial_excel(request):
    """Exportar resumen financiero a Excel."""
    today = timezone.now().date()
    
    # Obtener mes y año
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    
    # Calcular rango de fechas
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    # Obtener datos
    transactions = FinanceTransaction.objects.filter(
        transaction_date__gte=start_date,
        transaction_date__lte=end_date,
        is_active=True
    ).select_related('category').order_by('transaction_date')
    
    income_total = sum(t.amount_clp for t in transactions if t.category.type == 'income')
    expense_total = sum(t.amount_clp for t in transactions if t.category.type == 'expense')
    net_profit = income_total - expense_total
    
    # Calcular por categoría
    income_by_category = {}
    expense_by_category = {}
    
    for transaction in transactions:
        cat_name = transaction.category.category_name
        amount = transaction.amount_clp
        
        if transaction.category.type == 'income':
            income_by_category[cat_name] = income_by_category.get(cat_name, 0) + amount
        else:
            expense_by_category[cat_name] = expense_by_category.get(cat_name, 0) + amount
    
    # Crear workbook
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="1a237e")
    subtitle_font = Font(size=10, color="666666")
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========================================================================
    # HOJA 1: RESUMEN
    # ========================================================================
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    
    # Título
    month_names = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    ws_summary['A1'] = f"Resumen Financiero - {month_names[month]} {year}"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    # Información
    ws_summary['A2'] = "Avícola Eugenio"
    ws_summary['A2'].font = subtitle_font
    ws_summary['A3'] = f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    ws_summary['A3'].font = subtitle_font
    ws_summary['A4'] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws_summary['A4'].font = subtitle_font
    
    # Resumen financiero
    ws_summary['A6'] = "RESUMEN FINANCIERO"
    ws_summary['A6'].font = header_font
    ws_summary['A6'].fill = header_fill
    ws_summary['B6'] = "MONTO"
    ws_summary['B6'].font = header_font
    ws_summary['B6'].fill = header_fill
    ws_summary['B6'].alignment = Alignment(horizontal='right')
    
    ws_summary['A7'] = "Total Ingresos"
    ws_summary['B7'] = income_total
    ws_summary['B7'].number_format = '"$"#,##0'
    ws_summary['B7'].alignment = Alignment(horizontal='right')
    
    ws_summary['A8'] = "Total Egresos"
    ws_summary['B8'] = expense_total
    ws_summary['B8'].number_format = '"$"#,##0'
    ws_summary['B8'].alignment = Alignment(horizontal='right')
    
    ws_summary['A9'] = "Utilidad Neta"
    ws_summary['A9'].font = bold_font
    ws_summary['B9'] = net_profit
    ws_summary['B9'].number_format = '"$"#,##0'
    ws_summary['B9'].alignment = Alignment(horizontal='right')
    ws_summary['B9'].font = Font(bold=True, color="008000" if net_profit >= 0 else "FF0000")
    
    # Aplicar bordes
    for row in range(6, 10):
        for col in ['A', 'B']:
            ws_summary[f'{col}{row}'].border = border
    
    # Ajustar anchos de columna
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # Ingresos por categoría
    if income_by_category:
        row = 12
        ws_summary[f'A{row}'] = "INGRESOS POR CATEGORÍA"
        ws_summary[f'A{row}'].font = header_font
        ws_summary[f'A{row}'].fill = header_fill
        ws_summary[f'B{row}'] = "MONTO"
        ws_summary[f'B{row}'].font = header_font
        ws_summary[f'B{row}'].fill = header_fill
        ws_summary[f'B{row}'].alignment = Alignment(horizontal='right')
        
        row += 1
        for cat_name, amount in sorted(income_by_category.items()):
            ws_summary[f'A{row}'] = cat_name
            ws_summary[f'B{row}'] = amount
            ws_summary[f'B{row}'].number_format = '"$"#,##0'
            ws_summary[f'B{row}'].alignment = Alignment(horizontal='right')
            ws_summary[f'A{row}'].border = border
            ws_summary[f'B{row}'].border = border
            row += 1
    
    # Egresos por categoría
    if expense_by_category:
        row += 2
        ws_summary[f'A{row}'] = "EGRESOS POR CATEGORÍA"
        ws_summary[f'A{row}'].font = header_font
        ws_summary[f'A{row}'].fill = header_fill
        ws_summary[f'B{row}'] = "MONTO"
        ws_summary[f'B{row}'].font = header_font
        ws_summary[f'B{row}'].fill = header_fill
        ws_summary[f'B{row}'].alignment = Alignment(horizontal='right')
        
        row += 1
        for cat_name, amount in sorted(expense_by_category.items()):
            ws_summary[f'A{row}'] = cat_name
            ws_summary[f'B{row}'] = amount
            ws_summary[f'B{row}'].number_format = '"$"#,##0'
            ws_summary[f'B{row}'].alignment = Alignment(horizontal='right')
            ws_summary[f'A{row}'].border = border
            ws_summary[f'B{row}'].border = border
            row += 1
    
    # ========================================================================
    # HOJA 2: DETALLE DE TRANSACCIONES
    # ========================================================================
    ws_detail = wb.create_sheet("Detalle Transacciones")
    
    # Encabezados
    headers = ['Fecha', 'Categoría', 'Tipo', 'Monto', 'Método Pago', 'Referencia', 'Descripción']
    for col_num, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos
    for row_num, transaction in enumerate(transactions, 2):
        ws_detail.cell(row=row_num, column=1, value=transaction.transaction_date)
        ws_detail.cell(row=row_num, column=1).number_format = 'DD/MM/YYYY'
        ws_detail.cell(row=row_num, column=2, value=transaction.category.category_name)
        ws_detail.cell(row=row_num, column=3, value='Ingreso' if transaction.category.type == 'income' else 'Egreso')
        ws_detail.cell(row=row_num, column=4, value=float(transaction.amount_clp))
        ws_detail.cell(row=row_num, column=4).number_format = '"$"#,##0'
        ws_detail.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')
        ws_detail.cell(row=row_num, column=5, value=transaction.payment_method)
        ws_detail.cell(row=row_num, column=6, value=transaction.reference_doc)
        ws_detail.cell(row=row_num, column=7, value=transaction.description)
        
        # Aplicar bordes
        for col_num in range(1, 8):
            ws_detail.cell(row=row_num, column=col_num).border = border
    
    # Ajustar anchos de columna
    ws_detail.column_dimensions['A'].width = 12
    ws_detail.column_dimensions['B'].width = 25
    ws_detail.column_dimensions['C'].width = 10
    ws_detail.column_dimensions['D'].width = 15
    ws_detail.column_dimensions['E'].width = 15
    ws_detail.column_dimensions['F'].width = 15
    ws_detail.column_dimensions['G'].width = 40
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="resumen_financiero_{year}_{month:02d}.xlsx"'
    
    # Guardar workbook
    wb.save(response)
    
    return response
